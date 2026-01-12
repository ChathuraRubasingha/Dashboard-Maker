from typing import List, Optional, Dict, Any, Tuple
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
import secrets
import os
import re
import uuid
import io
from datetime import datetime

from app.models.excel_report import ExcelTemplateReport
from app.schemas.excel_report import (
    ExcelTemplateReportCreate,
    ExcelTemplateReportUpdate,
    ExcelPlaceholder,
    DataSourceMapping,
    PlaceholderType,
)

# Try to import openpyxl - it's optional for development
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Placeholder pattern: {{type:name}}
PLACEHOLDER_PATTERN = re.compile(r'\{\{(table|value|chart):(\w+)\}\}')


class ExcelReportService:
    """Service for managing Excel template reports."""

    def __init__(self, db: AsyncSession):
        self.db = db
        # Use absolute path for upload directory (works on Windows and Unix)
        self.upload_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads", "excel_templates")
        # Ensure upload directory exists
        os.makedirs(self.upload_dir, exist_ok=True)

    async def get_reports(self, include_archived: bool = False) -> List[ExcelTemplateReport]:
        """Get all Excel template reports."""
        query = select(ExcelTemplateReport)
        if not include_archived:
            query = query.where(ExcelTemplateReport.is_archived == False)
        query = query.order_by(
            ExcelTemplateReport.updated_at.desc().nullsfirst(),
            ExcelTemplateReport.created_at.desc()
        )
        result = await self.db.execute(query)
        return result.scalars().all()

    async def get_report(self, report_id: int) -> Optional[ExcelTemplateReport]:
        """Get a single Excel template report by ID."""
        query = select(ExcelTemplateReport).where(ExcelTemplateReport.id == report_id)
        result = await self.db.execute(query)
        return result.scalar_one_or_none()

    async def get_report_by_share_token(self, share_token: str) -> Optional[ExcelTemplateReport]:
        """Get a public Excel template report by its share token."""
        query = (
            select(ExcelTemplateReport)
            .where(ExcelTemplateReport.share_token == share_token)
            .where(ExcelTemplateReport.is_public == True)
        )
        result = await self.db.execute(query)
        return result.scalar_one_or_none()

    async def create_report(self, data: ExcelTemplateReportCreate) -> ExcelTemplateReport:
        """Create a new Excel template report."""
        report = ExcelTemplateReport(
            name=data.name,
            description=data.description,
            placeholders=[],
            mappings={},
        )
        self.db.add(report)
        await self.db.commit()
        await self.db.refresh(report)
        return report

    async def update_report(
        self, report_id: int, data: ExcelTemplateReportUpdate
    ) -> Optional[ExcelTemplateReport]:
        """Update an Excel template report."""
        report = await self.get_report(report_id)
        if not report:
            return None

        update_data = data.model_dump(exclude_unset=True)

        # Handle mappings conversion
        if "mappings" in update_data and update_data["mappings"]:
            mappings_dict = {}
            for key, mapping in update_data["mappings"].items():
                if hasattr(mapping, "model_dump"):
                    mappings_dict[key] = mapping.model_dump()
                else:
                    mappings_dict[key] = mapping
            update_data["mappings"] = mappings_dict

        for field, value in update_data.items():
            setattr(report, field, value)

        await self.db.commit()
        await self.db.refresh(report)
        return report

    async def delete_report(self, report_id: int) -> bool:
        """Delete an Excel template report and its template file."""
        report = await self.get_report(report_id)
        if not report:
            return False

        # Delete template file if exists
        if report.template_file_path and os.path.exists(report.template_file_path):
            try:
                os.remove(report.template_file_path)
            except OSError:
                pass  # Ignore file deletion errors

        await self.db.delete(report)
        await self.db.commit()
        return True

    async def upload_template(
        self, report_id: int, file_content: bytes, filename: str
    ) -> Tuple[List[ExcelPlaceholder], str]:
        """
        Upload a template file and detect placeholders.
        Returns tuple of (placeholders, saved_filename)
        """
        report = await self.get_report(report_id)
        if not report:
            raise ValueError("Report not found")

        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is not installed. Please install it with: pip install openpyxl")

        # Delete old template if exists
        if report.template_file_path and os.path.exists(report.template_file_path):
            try:
                os.remove(report.template_file_path)
            except OSError:
                pass

        # Save new template
        safe_filename = f"{report_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{filename}"
        file_path = os.path.join(self.upload_dir, safe_filename)

        with open(file_path, 'wb') as f:
            f.write(file_content)

        # Detect placeholders
        placeholders = self._detect_placeholders(file_content)

        # Update report
        report.template_file_path = file_path
        report.template_filename = filename
        report.placeholders = [p.model_dump() for p in placeholders]
        report.mappings = {}  # Reset mappings when template changes

        await self.db.commit()
        await self.db.refresh(report)

        return placeholders, filename

    def _detect_placeholders(self, file_content: bytes) -> List[ExcelPlaceholder]:
        """Detect all placeholders in an Excel template."""
        if not OPENPYXL_AVAILABLE:
            return []

        placeholders = []
        wb = load_workbook(io.BytesIO(file_content))

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        matches = PLACEHOLDER_PATTERN.findall(cell.value)
                        for match in matches:
                            placeholder_type, placeholder_name = match
                            placeholder = ExcelPlaceholder(
                                id=str(uuid.uuid4()),
                                placeholder=f"{{{{{placeholder_type}:{placeholder_name}}}}}",
                                type=PlaceholderType(placeholder_type),
                                name=placeholder_name,
                                sheet_name=sheet_name,
                                cell_reference=cell.coordinate,
                            )
                            placeholders.append(placeholder)

        wb.close()
        return placeholders

    async def update_mappings(
        self, report_id: int, mappings: Dict[str, DataSourceMapping]
    ) -> Optional[ExcelTemplateReport]:
        """Update placeholder to data source mappings."""
        report = await self.get_report(report_id)
        if not report:
            return None

        # Convert mappings to dict
        mappings_dict = {}
        for key, mapping in mappings.items():
            if hasattr(mapping, "model_dump"):
                mappings_dict[key] = mapping.model_dump()
            else:
                mappings_dict[key] = mapping

        report.mappings = mappings_dict

        await self.db.commit()
        await self.db.refresh(report)
        return report

    async def generate_excel(
        self, report_id: int, data_fetcher: callable
    ) -> Tuple[bytes, str]:
        """
        Generate an Excel file by filling the template with data.

        Args:
            report_id: ID of the Excel template report
            data_fetcher: Async function to fetch data for a mapping
                          Signature: async def fetch(mapping: dict) -> list[list]

        Returns:
            Tuple of (file_bytes, filename)
        """
        report = await self.get_report(report_id)
        if not report:
            raise ValueError("Report not found")

        if not report.template_file_path or not os.path.exists(report.template_file_path):
            raise ValueError("Template file not found")

        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is not installed")

        # Load template
        wb = load_workbook(report.template_file_path)

        # Process each placeholder
        for placeholder_data in report.placeholders:
            placeholder_id = placeholder_data.get('id')
            mapping = report.mappings.get(placeholder_id)

            if not mapping:
                continue

            # Fetch data for this placeholder
            try:
                data = await data_fetcher(mapping)
            except Exception as e:
                print(f"Error fetching data for placeholder {placeholder_id}: {e}")
                continue

            # Fill placeholder in template
            sheet_name = placeholder_data.get('sheet_name')
            cell_ref = placeholder_data.get('cell_reference')
            placeholder_type = placeholder_data.get('type')

            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]

            if placeholder_type == 'value':
                # Single value - replace the cell content
                if data and len(data) > 0 and len(data[0]) > 0:
                    ws[cell_ref] = data[0][0]
                else:
                    ws[cell_ref] = ""

            elif placeholder_type == 'table':
                # Table data - fill starting from the cell
                self._fill_table_data(ws, cell_ref, data)

            elif placeholder_type == 'chart':
                # Chart - for now just put a placeholder text
                ws[cell_ref] = "[Chart data]"

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)

        filename = f"{report.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return output.getvalue(), filename

    def _fill_table_data(self, ws, start_cell: str, data: List[List[Any]]):
        """Fill table data starting from the given cell."""
        if not data:
            ws[start_cell] = ""
            return

        # Parse start cell reference
        col_letter = ''.join(c for c in start_cell if c.isalpha())
        row_num = int(''.join(c for c in start_cell if c.isdigit()))

        # Clear the placeholder
        ws[start_cell] = ""

        # Write data
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                col = self._get_column_index(col_letter) + col_idx
                cell = ws.cell(row=row_num + row_idx, column=col)
                cell.value = value

    def _get_column_index(self, col_letter: str) -> int:
        """Convert column letter(s) to 1-based index."""
        result = 0
        for char in col_letter.upper():
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result

    async def generate_share_token(self, report_id: int) -> Optional[ExcelTemplateReport]:
        """Generate or regenerate a share token for a report."""
        report = await self.get_report(report_id)
        if not report:
            return None

        report.share_token = secrets.token_urlsafe(32)
        report.is_public = True

        await self.db.commit()
        await self.db.refresh(report)
        return report

    async def revoke_share(self, report_id: int) -> Optional[ExcelTemplateReport]:
        """Revoke sharing for a report."""
        report = await self.get_report(report_id)
        if not report:
            return None

        report.share_token = None
        report.is_public = False

        await self.db.commit()
        await self.db.refresh(report)
        return report

    async def duplicate_report(
        self, report_id: int, new_name: Optional[str] = None
    ) -> Optional[ExcelTemplateReport]:
        """Duplicate an Excel template report."""
        original = await self.get_report(report_id)
        if not original:
            return None

        # Copy template file if exists
        new_file_path = None
        if original.template_file_path and os.path.exists(original.template_file_path):
            new_filename = f"copy_{datetime.now().strftime('%Y%m%d%H%M%S')}_{original.template_filename}"
            new_file_path = os.path.join(self.upload_dir, new_filename)
            with open(original.template_file_path, 'rb') as src:
                with open(new_file_path, 'wb') as dst:
                    dst.write(src.read())

        duplicate = ExcelTemplateReport(
            name=new_name or f"{original.name} (Copy)",
            description=original.description,
            template_file_path=new_file_path,
            template_filename=original.template_filename,
            placeholders=original.placeholders.copy() if original.placeholders else [],
            mappings=original.mappings.copy() if original.mappings else {},
            is_public=False,
            share_token=None,
        )

        self.db.add(duplicate)
        await self.db.commit()
        await self.db.refresh(duplicate)
        return duplicate
