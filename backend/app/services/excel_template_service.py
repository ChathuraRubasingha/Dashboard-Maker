from typing import List, Optional, Dict, Any
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from fastapi import UploadFile
import os
import uuid
import secrets

from app.models.excel_template import ExcelTemplate
from app.models.excel_report import ExcelTemplateReport
from app.schemas.excel_template import (
    ExcelTemplateCreate,
    ExcelTemplateUpdate,
    ExcelReportCreate,
    ExcelReportUpdate,
    DataSourceMapping,
)

# Try to import openpyxl
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def rgb_to_hex(rgb_color) -> Optional[str]:
    """Convert openpyxl color to hex string."""
    if rgb_color is None:
        return None
    if hasattr(rgb_color, 'rgb') and rgb_color.rgb:
        rgb = rgb_color.rgb
        if isinstance(rgb, str) and len(rgb) >= 6:
            # Remove alpha channel if present (ARGB format)
            if len(rgb) == 8:
                return f"#{rgb[2:]}"
            return f"#{rgb}"
    return None


def parse_excel_template(file_path: str) -> Dict[str, Any]:
    """
    Parse an Excel file and extract its structure.
    Returns a dictionary with sheets, cells, styles, merges, etc.
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is not installed")

    wb = load_workbook(file_path, data_only=False)
    structure = {"sheets": []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = {
            "name": sheet_name,
            "cells": {},
            "merges": [],
            "columnWidths": {},
            "rowHeights": {},
        }

        # Parse merged cells
        for merge_range in ws.merged_cells.ranges:
            sheet_data["merges"].append(str(merge_range))

        # Parse column widths
        for col_idx, col_dim in ws.column_dimensions.items():
            if col_dim.width:
                # Convert column letter to index
                if isinstance(col_idx, str):
                    col_num = column_index_from_string(col_idx) - 1
                else:
                    col_num = col_idx - 1
                sheet_data["columnWidths"][col_num] = col_dim.width * 7  # Approximate pixels

        # Parse row heights
        for row_idx, row_dim in ws.row_dimensions.items():
            if row_dim.height:
                sheet_data["rowHeights"][row_idx - 1] = row_dim.height

        # Parse cells
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None or cell.font or cell.fill or cell.border or cell.alignment:
                    cell_ref = cell.coordinate
                    cell_data = {}

                    # Value
                    if cell.value is not None:
                        cell_data["value"] = cell.value
                        if isinstance(cell.value, (int, float)):
                            cell_data["value"] = cell.value
                        else:
                            cell_data["value"] = str(cell.value)

                    # Formula
                    if cell.data_type == 'f' and cell.value:
                        cell_data["formula"] = str(cell.value)

                    # Style
                    style = {}

                    # Font
                    if cell.font:
                        font = {}
                        if cell.font.name:
                            font["name"] = cell.font.name
                        if cell.font.size:
                            font["size"] = cell.font.size
                        if cell.font.bold:
                            font["bold"] = cell.font.bold
                        if cell.font.italic:
                            font["italic"] = cell.font.italic
                        if cell.font.color:
                            color = rgb_to_hex(cell.font.color)
                            if color:
                                font["color"] = color
                        if font:
                            style["font"] = font

                    # Fill
                    if cell.fill and cell.fill.fgColor:
                        color = rgb_to_hex(cell.fill.fgColor)
                        if color and color != "#000000":
                            style["fill"] = {"color": color}

                    # Border
                    if cell.border:
                        border = {}
                        for side in ["top", "bottom", "left", "right"]:
                            side_obj = getattr(cell.border, side)
                            if side_obj and side_obj.style:
                                border[side] = {
                                    "style": side_obj.style,
                                    "color": rgb_to_hex(side_obj.color) if side_obj.color else None
                                }
                        if border:
                            style["border"] = border

                    # Alignment
                    if cell.alignment:
                        alignment = {}
                        if cell.alignment.horizontal:
                            alignment["horizontal"] = cell.alignment.horizontal
                        if cell.alignment.vertical:
                            alignment["vertical"] = cell.alignment.vertical
                        if cell.alignment.wrap_text:
                            alignment["wrapText"] = cell.alignment.wrap_text
                        if alignment:
                            style["alignment"] = alignment

                    # Number format
                    if cell.number_format and cell.number_format != "General":
                        style["numberFormat"] = cell.number_format

                    if style:
                        cell_data["style"] = style

                    if cell_data:
                        sheet_data["cells"][cell_ref] = cell_data

        structure["sheets"].append(sheet_data)

    wb.close()
    return structure


def generate_excel_with_data(
    template_path: str,
    sheet_data: Dict[str, Any],
    data_sources: List[Dict[str, Any]],
    visualization_data: Dict[int, List[Dict[str, Any]]]
) -> bytes:
    """
    Generate an Excel file from a template with data filled in.

    Args:
        template_path: Path to the template file
        sheet_data: Cell modifications made by user
        data_sources: Data source mappings
        visualization_data: Dict of visualization_id -> data rows

    Returns:
        Excel file as bytes
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is not installed")

    wb = load_workbook(template_path)

    # Apply cell modifications
    if sheet_data and "sheets" in sheet_data:
        for sheet_info in sheet_data["sheets"]:
            sheet_name = sheet_info.get("name")
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                cells = sheet_info.get("cells", {})
                for cell_ref, cell_data in cells.items():
                    ws[cell_ref] = cell_data.get("value", "")

    # Apply data source mappings
    for mapping in data_sources:
        viz_id = mapping.get("visualization_id")
        sheet_name = mapping.get("sheet_name")
        start_cell = mapping.get("start_cell", "A1")
        columns = mapping.get("columns", [])
        include_header = mapping.get("include_header", True)

        if viz_id not in visualization_data:
            continue

        data_rows = visualization_data[viz_id]
        if not data_rows:
            continue

        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        # Parse start cell
        start_col = ""
        start_row = ""
        for char in start_cell:
            if char.isalpha():
                start_col += char
            else:
                start_row += char

        start_col_idx = column_index_from_string(start_col)
        start_row_idx = int(start_row)

        current_row = start_row_idx

        # Write header if requested
        if include_header and columns:
            for col_offset, col_map in enumerate(columns):
                cell = ws.cell(row=current_row, column=start_col_idx + col_offset)
                # Use header_label (custom name) if available, otherwise fall back to source_column
                cell.value = col_map.get("header_label") or col_map.get("source_column", "")
                # Style header
                cell.font = Font(bold=True)
            current_row += 1

        # Write data rows
        for data_row in data_rows:
            for col_offset, col_map in enumerate(columns):
                source_col = col_map.get("source_column")
                cell = ws.cell(row=current_row, column=start_col_idx + col_offset)
                cell.value = data_row.get(source_col, "")
            current_row += 1

    # Save to bytes
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)
    return output.read()


class ExcelTemplateService:
    """Service for managing Excel templates."""

    def __init__(self, db: AsyncSession):
        self.db = db
        self.upload_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            "uploads", "excel_templates"
        )
        os.makedirs(self.upload_dir, exist_ok=True)

    async def get_templates(self, include_archived: bool = False) -> List[ExcelTemplate]:
        """Get all Excel templates."""
        query = select(ExcelTemplate)
        if not include_archived:
            query = query.where(ExcelTemplate.is_archived == False)
        query = query.order_by(
            ExcelTemplate.updated_at.desc().nullsfirst(),
            ExcelTemplate.created_at.desc()
        )
        result = await self.db.execute(query)
        return result.scalars().all()

    async def get_template(self, template_id: int) -> Optional[ExcelTemplate]:
        """Get a single template by ID."""
        query = select(ExcelTemplate).where(ExcelTemplate.id == template_id)
        result = await self.db.execute(query)
        return result.scalar_one_or_none()

    async def create_template(self, data: ExcelTemplateCreate) -> ExcelTemplate:
        """Create a new template (without file)."""
        template = ExcelTemplate(
            name=data.name,
            description=data.description,
            structure={},
        )
        self.db.add(template)
        await self.db.commit()
        await self.db.refresh(template)
        return template

    async def upload_template_file(
        self,
        template_id: int,
        file: UploadFile
    ) -> Optional[ExcelTemplate]:
        """Upload and parse a template file."""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is not installed")

        template = await self.get_template(template_id)
        if not template:
            return None

        # Save file
        file_ext = os.path.splitext(file.filename)[1] if file.filename else ".xlsx"
        unique_filename = f"{uuid.uuid4()}{file_ext}"
        file_path = os.path.join(self.upload_dir, unique_filename)

        content = await file.read()
        with open(file_path, "wb") as f:
            f.write(content)

        # Parse template
        try:
            structure = parse_excel_template(file_path)
        except Exception as e:
            # Clean up file on error
            if os.path.exists(file_path):
                os.remove(file_path)
            raise e

        # Update template
        template.file_path = file_path
        template.file_name = file.filename
        template.structure = structure

        await self.db.commit()
        await self.db.refresh(template)
        return template

    async def update_template(
        self,
        template_id: int,
        data: ExcelTemplateUpdate
    ) -> Optional[ExcelTemplate]:
        """Update a template."""
        template = await self.get_template(template_id)
        if not template:
            return None

        update_data = data.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(template, field, value)

        await self.db.commit()
        await self.db.refresh(template)
        return template

    async def delete_template(self, template_id: int) -> bool:
        """Delete a template."""
        template = await self.get_template(template_id)
        if not template:
            return False

        # Delete file if exists
        if template.file_path and os.path.exists(template.file_path):
            os.remove(template.file_path)

        await self.db.delete(template)
        await self.db.commit()
        return True


class ExcelReportService:
    """Service for managing Excel reports."""

    def __init__(self, db: AsyncSession):
        self.db = db

    async def get_reports(self, include_archived: bool = False) -> List[ExcelTemplateReport]:
        """Get all Excel reports."""
        query = select(ExcelTemplateReport)
        if not include_archived:
            query = query.where(ExcelTemplateReport.is_archived == False)
        query = query.order_by(
            ExcelTemplateReport.updated_at.desc().nullsfirst(),
            ExcelTemplateReport.created_at.desc()
        )
        result = await self.db.execute(query)
        return result.scalars().all()

    async def get_report(self, report_id: int) -> Optional[ExcelTemplateReport]:
        """Get a single report by ID."""
        query = select(ExcelTemplateReport).where(ExcelTemplateReport.id == report_id)
        result = await self.db.execute(query)
        return result.scalar_one_or_none()

    async def get_report_by_share_token(self, share_token: str) -> Optional[ExcelTemplateReport]:
        """Get a public report by share token."""
        query = (
            select(ExcelTemplateReport)
            .where(ExcelTemplateReport.share_token == share_token)
            .where(ExcelTemplateReport.is_public == True)
        )
        result = await self.db.execute(query)
        return result.scalar_one_or_none()

    async def create_report(self, data: ExcelReportCreate) -> ExcelTemplateReport:
        """Create a new report from template."""
        # Convert data_sources to dict if present
        data_sources_list = []
        if data.data_sources:
            data_sources_list = [
                ds.model_dump() if hasattr(ds, "model_dump") else ds
                for ds in data.data_sources
            ]

        report = ExcelTemplateReport(
            name=data.name,
            description=data.description,
            template_id=data.template_id,
            sheet_data={},
            data_sources=data_sources_list,
        )
        self.db.add(report)
        await self.db.commit()
        await self.db.refresh(report)
        return report

    async def update_report(
        self,
        report_id: int,
        data: ExcelReportUpdate
    ) -> Optional[ExcelTemplateReport]:
        """Update a report."""
        report = await self.get_report(report_id)
        if not report:
            return None

        update_data = data.model_dump(exclude_unset=True)

        # Convert data_sources to dict if present
        if "data_sources" in update_data and update_data["data_sources"]:
            update_data["data_sources"] = [
                ds.model_dump() if hasattr(ds, "model_dump") else ds
                for ds in update_data["data_sources"]
            ]

        for field, value in update_data.items():
            setattr(report, field, value)

        await self.db.commit()
        await self.db.refresh(report)
        return report

    async def delete_report(self, report_id: int) -> bool:
        """Delete a report."""
        report = await self.get_report(report_id)
        if not report:
            return False

        await self.db.delete(report)
        await self.db.commit()
        return True

    async def generate_share_token(self, report_id: int) -> Optional[ExcelTemplateReport]:
        """Generate or regenerate a share token."""
        report = await self.get_report(report_id)
        if not report:
            return None

        report.share_token = secrets.token_urlsafe(32)
        report.is_public = True

        await self.db.commit()
        await self.db.refresh(report)
        return report

    async def revoke_share(self, report_id: int) -> Optional[ExcelTemplateReport]:
        """Revoke sharing for a report."""
        report = await self.get_report(report_id)
        if not report:
            return None

        report.share_token = None
        report.is_public = False

        await self.db.commit()
        await self.db.refresh(report)
        return report

    async def generate_excel(
        self,
        report_id: int,
        visualization_data: Dict[int, List[Dict[str, Any]]]
    ) -> Optional[bytes]:
        """Generate Excel file with data."""
        report = await self.get_report(report_id)
        if not report or not report.template_id:
            return None

        # Get template
        template_query = select(ExcelTemplate).where(ExcelTemplate.id == report.template_id)
        result = await self.db.execute(template_query)
        template = result.scalar_one_or_none()

        if not template or not template.file_path:
            return None

        return generate_excel_with_data(
            template.file_path,
            report.sheet_data or {},
            report.data_sources or [],
            visualization_data
        )
